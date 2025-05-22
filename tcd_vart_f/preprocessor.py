import streamlit as st
import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side
import io

def load_and_preprocess_tcd(df):
    df = df[["Labels", "Action", "Expected Results"]]

    df["Action"] = df["Action"].fillna("").astype(str)
    df["Expected Results"] = df["Expected Results"].fillna("").astype(str)
    df["Test_Case_Type"] = df["Labels"].apply(lambda x: x.split("_")[-1].strip().lower().replace(" ", "") if isinstance(x, str) else "")
    df["Sub_Feature"] = df["Labels"].apply(lambda x: "_".join(x.split("_")[2:-1]) if isinstance(x, str) else "")
    df["Normalized_Feature"] = df["Sub_Feature"].apply(lambda x: re.sub(r"[_\s]+", "_", x.strip().lower()) if isinstance(x, str) else "")
    return df


def extract_steps(row):
    action_steps = []
    expected_steps = {}


    action_content = str(row["Action"]) if pd.notna(row["Action"]) else ""
    expected_content = str(row["Expected Results"]) if pd.notna(row["Expected Results"]) else ""

    action_lines = action_content.split("\n")
    extracting = False
    for line in action_lines:
        if "Steps:" in line:
            extracting = True
            continue
        if extracting and line.strip():
            clean_line = re.sub(r"^\d+\.\s*", "", line.strip())
            action_steps.append(clean_line)

    expected_lines = expected_content.split("\n")
    for line in expected_lines:
        if line.strip() and re.match(r"^\d+\.", line):
            parts = line.split(".", 1)
            try:
                step_num = int(parts[0].strip())
                clean_value = parts[1].strip()
                expected_steps[step_num] = clean_value
            except ValueError:
                continue

    final_steps = []
    step_counter = 1
    used_expected = set()
    battery_reconnect_steps = ["RELAY_OFF: ign", "RELAY_OFF: bat", "RELAY_ON: bat", "RELAY_ON: ign"]

    for i, action in enumerate(action_steps, start=1):
        if "battery reconnect" in action.lower():
            for step in battery_reconnect_steps:
                final_steps.append(f"{step_counter}. {step}")
                step_counter += 1
        else:
            final_steps.append(f"{step_counter}. {action}")

        if i in expected_steps:
            final_steps.append(f"{step_counter + 1}. {expected_steps[i]}")
            used_expected.add(i)
            step_counter += 1
        step_counter += 1

    for key in sorted(expected_steps.keys()):
        if key not in used_expected:
            final_steps.append(f"{step_counter}. {expected_steps[key]}")
            step_counter += 1

    return "\n".join(final_steps)

def generate_vart_sheet(df):
    default_keywords = ["TC_ID", "RELAY_ON", "RELAY_ON", "WAIT_S", "INIT", "WAIT_S"]
    default_values = ["", "bat", "ign", "5", "yes", "5"]
    vart_data = []
    category_order = ["logicalcombination", "failuremodes", "powermodes", "configuration", "voltagemodes"]
    category_names = {"logicalcombination": "Logical", "failuremodes": "Failure modes", "powermodes": "Power modes", "configuration": "Configuration modes", "voltagemodes": "Voltage modes"}
    feature_order = df.drop_duplicates(subset="Normalized_Feature")[["Normalized_Feature", "Sub_Feature"]].values

    for normalized_feature, original_feature in feature_order:
        feature_df = df[df["Normalized_Feature"] == normalized_feature]
        vart_data.append([original_feature])

        for category in category_order:
            category_df = feature_df[feature_df["Test_Case_Type"] == category]
            if category_df.empty:
                continue

            vart_data.append([category_names[category]])

            for _, row in category_df.iterrows():
                
                steps_content = str(row["Steps"]) if pd.notna(row["Steps"]) else ""
                steps = steps_content.split("\n")
                
                step_keywords = []
                step_values = []

                for step in steps:
                    if ":" in step:
                        key, value = step.split(":", 1)
                        step_keywords.append(re.sub(r"^\d+\.\s*", "", key.strip()))
                        step_values.append(value.strip())

            
                vart_keywords = default_keywords + step_keywords + ["END"]
                vart_values = default_values + step_values + ["yes"]

                vart_data.append(vart_keywords)
                vart_data.append(vart_values)
                vart_data.append([""] * len(vart_keywords)) 

    wb = Workbook()
    ws = wb.active
    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    category_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    feature_fill = PatternFill(start_color="A9D08E", end_color="A9D08E", fill_type="solid")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    for row_idx, row in enumerate(vart_data, start=1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if any(row):
                cell.border = thin_border
            if len(row) == 1 and row[0] in df["Sub_Feature"].unique():
                cell.fill = feature_fill
            elif len(row) == 1 and row[0] in category_names.values():
                cell.fill = category_fill
            elif len(row) > 1 and row[0] in default_keywords + ["END"]:
                cell.fill = header_fill
    return wb, vart_data

# Streamlit App
st.title("TCD to VART Converter")

uploaded_file = st.file_uploader("Upload your TCD Excel file", type=["xlsx"])

if uploaded_file:
    try:
        df = pd.read_excel(uploaded_file)
        st.write("Original Data:")
        st.dataframe(df.head())


        df_processed = load_and_preprocess_tcd(df.copy())


        df_processed["Steps"] = df_processed.apply(extract_steps, axis=1)

        st.write("Processed Data with Extracted Steps:")
        st.dataframe(df_processed.head())

        
        wb, vart_data = generate_vart_sheet(df_processed.copy())

    
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        st.download_button(
            label="Download VART Excel File",
            data=output,
            file_name="VART_Output.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        st.error(f"An error occurred: {e}")